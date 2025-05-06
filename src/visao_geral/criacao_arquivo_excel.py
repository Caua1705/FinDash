from pathlib import Path
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill
from openpyxl.chart import BarChart, Reference
import streamlit as st
import pandas as pd

def formatar_arquivo_excel(sheet) -> None:
    if sheet.max_column<3:
        tipo="incompleto"
    else:
        tipo="completo"
    #Formatar Colunas:
    for colunas in sheet.columns:
        for celula in colunas:
            if isinstance(celula.value,(int,float)):
                celula.number_format='"R$"#,##0.00'
            letras_colunas=celula.column_letter
            sheet.column_dimensions[letras_colunas].width=20
            celula.border=Border(left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin"))
            celula.font=Font(name="Calibri")
    #Formatar Cabeçalho:
    for celula in sheet[1]:
        celula.alignment=Alignment(horizontal="center", vertical="center",)
        celula.fill=PatternFill(start_color="004080",end_color="004080",fill_type="solid")
        celula.font=Font(bold=True,color="FFFFFF")
    #Formatar Totais:
    for linha in sheet.iter_rows(min_row=sheet.max_row,max_row=sheet.max_row):
        for celula in linha:
            celula.font=Font(bold=True)
            celula.fill=PatternFill(start_color="FFFFF99C",end_color="FFFFF99C",fill_type="solid")

    grafico=BarChart()
    categorias=Reference(sheet,min_col=1,max_col=1,min_row=2,max_row=sheet.max_row-1)

    if tipo=="completo":
        #Formatar Categoria:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=1,max_col=1):
            for celula in linha:
                celula.fill=PatternFill(start_color="FFF4E6",end_color="FFF4E6",fill_type="solid")
                celula.alignment=Alignment(horizontal="left")
        #Formatar Receitas:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=2,max_col=2):
            for celula in linha:
                celula.fill=PatternFill(start_color="A9D08E",end_color="A9D08E",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")
        #Formatar Despesas:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=3,max_col=3):
            for celula in linha:
                celula.fill=PatternFill(start_color="FF5757",end_color="FF5757",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")
        dados=Reference(sheet,min_col=2,max_col=3,min_row=1,max_row=sheet.max_row-1)
        grafico.title="Receitas e Despesas por Categoria"

    else:
        #Formatar Categoria:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=1,max_col=1):
            for celula in linha:
                celula.fill=PatternFill(start_color="FFF4E6",end_color="FFF4E6",fill_type="solid")
                celula.alignment=Alignment(horizontal="left")
        #Formatar Valor Total:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=2,max_col=2):
            for celula in linha:
                celula.fill=PatternFill(start_color="A9D08E",end_color="A9D08E",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")

        dados=Reference(sheet,min_col=2,max_col=2,min_row=1,max_row=sheet.max_row-1)
        grafico.title="Receitas Mensais"
        
    grafico.add_data(dados,titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.y_axis.majorGridlines = None
    sheet.add_chart(grafico,"G1")

def gerar_arquivo_excel(df_receitas_despesas,df_receitas_mensais,data_referencia) -> None:
    with tempfile.TemporaryDirectory() as dir_temp:
        nome_arquivo=f"Relatório mensal - {data_referencia}.xlsx"
        diretorio_arquivo_temporario= Path(dir_temp) / nome_arquivo
        with pd.ExcelWriter(diretorio_arquivo_temporario) as escritor:
            df_receitas_despesas.to_excel(escritor,sheet_name="Receitas e Despesas",index=False)
            df_receitas_mensais.to_excel(escritor,sheet_name="Receitas Mensais",index=False)
            
        wb=load_workbook(diretorio_arquivo_temporario)
        sheet1=wb["Receitas e Despesas"]
        sheet2=wb["Receitas Mensais"]
        formatar_arquivo_excel(sheet1)
        formatar_arquivo_excel(sheet2)
        wb.save(diretorio_arquivo_temporario)

        with open(diretorio_arquivo_temporario,"rb") as leitor:
            arquivo=leitor.read()
            st.download_button(" 📥 Clique para fazer o download",
                                        data=arquivo,
                                        file_name=nome_arquivo,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.divider()
            st.caption("Desenvolvido por Cauã de Carvalho Oliveira Peixoto - Todos os direitos reservados © 2025")

