import pandas as pd
import streamlit as st
import plotly.express as px
from pathlib import Path
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill
from openpyxl.chart import BarChart, Reference

def carregar_arquivo():
    st.set_page_config(page_title="FinDash",layout="wide",page_icon="💲") 
    st.title("Dashboard Financeiro:")
    st.write("Escolha um arquivo do tipo CSV ou XLSX para carregar a planilha:")
    upload_planilha=st.file_uploader("Selecione o arquivo:",accept_multiple_files=False,type=["xlsx","csv"])
    return upload_planilha

def carregar_dataframe(arquivo) -> pd.DataFrame:
    if arquivo.name.endswith("xlsx"):
        df=pd.read_excel(arquivo)
    if arquivo.name.endswith("csv"):
        df=pd.read_csv(arquivo)
    return df

def selecionar_colunas_dataframe(df) -> dict[str,pd.DataFrame]:
    col1,col2,col3=st.columns(3)
    with col1:
        coluna_data=st.selectbox("Selecione a coluna Data(dd/mm/yyyy)",list(df.columns),help="Coluna onde está a data da transação")
    with col2:
        coluna_centro_custo=st.selectbox("Selecione a coluna Centro de Custo",[a for a in list(df.columns) if a != coluna_data],help="Coluna onde está o centro de custo da transação")
    with col3:
        coluna_valor=st.selectbox("Selecione a coluna Valor",[a for a in list(df.columns) if a != coluna_data and a != coluna_centro_custo],help="Coluna onde está o valor da transação")
    dict_colunas={"Data":coluna_data,"Centro de Custo":coluna_centro_custo,"Valor":coluna_valor}
    return dict_colunas
        
def formatar_colunas_dataframe(df,colunas_dataframe) -> pd.DataFrame:
    df_formatado=df.copy()
    df_formatado=df_formatado.rename(columns={colunas_dataframe["Data"] :"Data",
                                      colunas_dataframe["Centro de Custo"] :"Centro de Custo",
                                      colunas_dataframe["Valor"] : "Valor"
                                      })
    df_formatado=df_formatado[["Data","Centro de Custo","Valor"]]   
    df_formatado["Centro de Custo"]=df_formatado["Centro de Custo"].fillna("Desconhecido")
    try:
        coluna_formatada=pd.to_datetime(df_formatado["Data"],dayfirst=True) 
    except ValueError:
           st.error(f"Erro na coluna Data. Revise os dados da tabela")
           st.stop()
    if df_formatado["Centro de Custo"].apply(lambda x: not isinstance(x,str)).any():
        st.error(f"Erro na coluna Centro de Custo. Revise os dados da tabela")
        st.stop()
    try:
        df_formatado["Tipo"]=df_formatado["Valor"].apply(lambda x: "receita" if x > 0 else "despesa")
    except TypeError:
        st.error(f"Erro na coluna Valor(R$). Revise os dados da tabela")
        st.stop()
    df_formatado["Data"]=coluna_formatada
    df_formatado["Valor"]=df_formatado["Valor"].apply(abs)
    return df_formatado

def filtrar_df_formatado_por_ano_mes(df_formatado) -> tuple[pd.DataFrame,str,str]:
    col1,col2=st.columns(2)
    numero_para_meses={1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}
    meses_para_numero={v:k for k, v in numero_para_meses.items() }
    ano=df_formatado["Data"].dt.year.unique()
    with col1:  
        filtro_ano=st.selectbox("Selecione o ano:",ano)
        meses_disponiveis=df_formatado.loc[df_formatado["Data"].dt.year == filtro_ano,"Data"].dt.month.unique()
    with col2:  
        filtro_mes=st.selectbox("Selecione o mês:",[numero_para_meses[mes] for mes in meses_disponiveis])
        df_filtrado=df_formatado.loc[(df_formatado["Data"].dt.year == filtro_ano) &
                                     (df_formatado["Data"].dt.month == meses_para_numero[filtro_mes])]
        data_filtrada=df_filtrado["Data"].iloc[0]
        data_referencia=data_filtrada.strftime("%Y-%m")
    return df_filtrado,filtro_mes,data_referencia 

def filtrar_dataframes_para_graficos(df_filtrado) -> tuple[pd.DataFrame,pd.DataFrame]:
    df_receitas_despesas=df_filtrado.pivot_table(index="Centro de Custo",
                                                       columns="Tipo",
                                                       values="Valor",
                                                       aggfunc="sum",
                                                       fill_value=0,
                                                       )
    df_receitas_despesas=df_receitas_despesas.rename(columns={"receita":"Receitas",
                                                                "despesa":"Despesas"}).reset_index()
    df_receitas_despesas=df_receitas_despesas[["Centro de Custo","Receitas","Despesas"]]
    df_receitas_despesas=df_receitas_despesas.sort_values(by="Receitas",ascending=False)
    df_receitas_despesas.loc[len(df_receitas_despesas)] = ["TOTAL",df_receitas_despesas["Receitas"].sum(),df_receitas_despesas["Despesas"].sum()]
    receitas_mensais=df_filtrado.loc[df_filtrado["Tipo"]=="receita"]
    df_receitas_mensais=receitas_mensais.groupby("Centro de Custo")["Valor"].sum().sort_values(ascending=False).reset_index()
    df_receitas_mensais.loc[len(df_receitas_mensais)] = ["TOTAL",df_receitas_mensais["Valor"].sum()]
    return df_receitas_despesas,df_receitas_mensais

def gerar_graficos(df_receitas_despesas,df_receitas_mensais,filtro_mes) -> None:
    col1,col2=st.columns(2)
    with col1:    
        st.subheader("Total de Receitas e Despesas")
        fig1=px.bar(df_receitas_despesas,x="Centro de Custo",y=["Receitas","Despesas"],barmode="group",labels={"Categoria": "Categoria", "valor": "Valor"},title=f"Receitas e Despesas por Categoria em {filtro_mes}")
        fig1.update_layout(xaxis_tickangle=-45,xaxis_title="Categoria",yaxis_title="Valor",showlegend=True)
        col1.plotly_chart(fig1)
    with col2:               
        st.subheader("Categorias com maiores Receitas")
        if len(df_receitas_mensais)>2:
            df_receitas_mensais=df_receitas_mensais.loc[0:2]
        fig2=px.pie(df_receitas_mensais,names="Centro de Custo",values="Valor",title=f"Distribuição das maiores Receitas em {filtro_mes}",color="Centro de Custo")
        fig2.update_traces(textinfo="percent+label")       
        col2.plotly_chart(fig2)
    st.divider()

def formatar_arquivo_excel(sheet) -> None:
    if sheet.max_column<3:
        tipo="incompleto"
    else:
        tipo="completo"
    #Formatar Colunas:
    for colunas in sheet.columns:
        for celula in colunas:
            if isinstance(celula.value,(int,float)):
                celula.number_format='"R$"#,##0.00'
            letras_colunas=celula.column_letter
            sheet.column_dimensions[letras_colunas].width=20
            celula.border=Border(left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin"))
            celula.font=Font(name="Calibri")
    #Formatar Cabeçalho:
    for celula in sheet[1]:
        celula.alignment=Alignment(horizontal="center", vertical="center",)
        celula.fill=PatternFill(start_color="004080",end_color="004080",fill_type="solid")
        celula.font=Font(bold=True,color="FFFFFF")
    #Formatar Totais:
    for linha in sheet.iter_rows(min_row=sheet.max_row,max_row=sheet.max_row):
        for celula in linha:
            celula.font=Font(bold=True)
            celula.fill=PatternFill(start_color="FFFFF99C",end_color="FFFFF99C",fill_type="solid")

    grafico=BarChart()
    categorias=Reference(sheet,min_col=1,max_col=1,min_row=2,max_row=sheet.max_row-1)

    if tipo=="completo":
        #Formatar Categoria:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=1,max_col=1):
            for celula in linha:
                celula.fill=PatternFill(start_color="FFF4E6",end_color="FFF4E6",fill_type="solid")
                celula.alignment=Alignment(horizontal="left")
        #Formatar Receitas:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=2,max_col=2):
            for celula in linha:
                celula.fill=PatternFill(start_color="A9D08E",end_color="A9D08E",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")
        #Formatar Despesas:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=3,max_col=3):
            for celula in linha:
                celula.fill=PatternFill(start_color="FF5757",end_color="FF5757",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")
        dados=Reference(sheet,min_col=2,max_col=3,min_row=1,max_row=sheet.max_row-1)
        grafico.title="Receitas e Despesas por Categoria"

    else:
        #Formatar Categoria:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=1,max_col=1):
            for celula in linha:
                celula.fill=PatternFill(start_color="FFF4E6",end_color="FFF4E6",fill_type="solid")
                celula.alignment=Alignment(horizontal="left")
        #Formatar Valor Total:
        for linha in sheet.iter_rows(min_row=2,max_row=sheet.max_row-1,min_col=2,max_col=2):
            for celula in linha:
                celula.fill=PatternFill(start_color="A9D08E",end_color="A9D08E",fill_type="solid")
                celula.alignment=Alignment(horizontal="right")

        dados=Reference(sheet,min_col=2,max_col=2,min_row=1,max_row=sheet.max_row-1)
        grafico.title="Receitas Mensais"
        
    grafico.add_data(dados,titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.y_axis.majorGridlines = None
    sheet.add_chart(grafico,"G1")

def criando_arquivo_excel(df_receitas_despesas,df_receitas_mensais,data_referencia) -> None:
    with tempfile.TemporaryDirectory() as dir_temp:
        nome_arquivo=f"Relatório mensal - {data_referencia}.xlsx"
        diretorio_arquivo_temporario= Path(dir_temp) / nome_arquivo
        with pd.ExcelWriter(diretorio_arquivo_temporario) as escritor:
            df_receitas_despesas.to_excel(escritor,sheet_name="Receitas e Despesas",index=False)
            df_receitas_mensais.to_excel(escritor,sheet_name="Receitas Mensais",index=False)
            
        wb=load_workbook(diretorio_arquivo_temporario)
        sheet1=wb["Receitas e Despesas"]
        sheet2=wb["Receitas Mensais"]
        formatar_arquivo_excel(sheet1)
        formatar_arquivo_excel(sheet2)
        wb.save(diretorio_arquivo_temporario)

        with open(diretorio_arquivo_temporario,"rb") as leitor:
            arquivo=leitor.read()
            st.markdown("### 📄 Exportação do relatório")
            st.markdown("Você pode baixar o relatório financeiro mensal em formato Excel.")
            st.download_button(" 📥 Clique para fazer o download",
                                        data=arquivo,
                                        file_name=nome_arquivo,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.divider()
            st.caption("Desenvolvido por Cauã de Carvalho Oliveira Peixoto - Todos os direitos reservados © 2025")
    
def main() -> None:
    upload_planilha=carregar_arquivo()
    if upload_planilha is not None:
        df=carregar_dataframe(upload_planilha)
        dict_colunas=selecionar_colunas_dataframe(df)
        df_formatado=formatar_colunas_dataframe(df,dict_colunas)
        df_filtrado,filtro_mes,data_referencia=filtrar_df_formatado_por_ano_mes(df_formatado)
        df_receitas_despesas,df_receitas_mensais=filtrar_dataframes_para_graficos(df_filtrado)
        gerar_graficos(df_receitas_despesas,df_receitas_mensais,filtro_mes)
        criando_arquivo_excel(df_receitas_despesas,df_receitas_mensais,data_referencia)
        
if __name__ == "__main__":
    main()
